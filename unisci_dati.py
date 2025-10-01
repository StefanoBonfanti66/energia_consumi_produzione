
import pandas as pd
import os
import re
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side # Added Border, Side
# from openpyxl.worksheet.properties import PageSetup, PrintPageSetup # Removed problematic import

try:
    # --- Nomi dei file ---
    file_quantita = 'prod_quantita.xlsx'
    file_consumo = 'prod_consumo_macchine.xlsx'
    file_output = 'report.xlsx'

    # --- Verifica esistenza file ---
    if not os.path.exists(file_quantita):
        print(f"Errore: Il file '{file_quantita}' non è stato trovato.")
        exit()
    if not os.path.exists(file_consumo):
        print(f"Errore: Il file '{file_consumo}' non è stato trovato.")
        exit()

    # --- Lettura File Quantità ---
    print(f"Leggo e trasformo il file: {file_quantita}")
    df_quantita = pd.read_excel(file_quantita, header=0)
    df_quantita[df_quantita.columns[0]] = df_quantita[df_quantita.columns[0]].astype(str)
    df_quantita_long = df_quantita.melt(
        id_vars=[df_quantita.columns[0]],
        var_name='Macchina',
        value_name='Quantita Prodotta'
    )
    df_quantita_long.rename(columns={df_quantita_long.columns[0]: 'Mese'}, inplace=True)

    # --- Lettura File Consumo ---
    print(f"Leggo e trasformo il file: {file_consumo}")
    df_consumo = pd.read_excel(file_consumo, header=2)
    date_column_name = df_consumo.columns[0]
    print(f"Colonna data/mese identificata come: '{date_column_name}'")
    df_consumo[date_column_name] = df_consumo[date_column_name].astype(str)
    df_consumo_long = df_consumo.melt(
        id_vars=[date_column_name],
        var_name='Macchina',
        value_name='Consumo Energia'
    )
    df_consumo_long.rename(columns={date_column_name: 'Mese'}, inplace=True)

    # --- Unione dei dati ---
    print("Unisco i dati...")
    df_merged = pd.merge(
        df_quantita_long,
        df_consumo_long,
        on=['Mese', 'Macchina'],
        how='outer'
    )

    # --- AGGREGAZIONE MENSILE ---
    print("Aggrego i dati su base mensile...")
    month_map = {
        'GENNAIO': '01', 'FEBBRAIO': '02', 'MARZO': '03', 'APRILE': '04',
        'MAGGIO': '05', 'GIUGNO': '06', 'GUIGNO': '06', 'LUGLIO': '07', 'AGOSTO': '08',
        'SETTEMBRE': '09', 'OTTOBRE': '10', 'NOVEMBRE': '11', 'DICEMBRE': '12'
    }

    def clean_and_parse_date(date_str):
        s = str(date_str).upper()
        for month_name, month_num in month_map.items():
            if month_name in s:
                match = re.search(r'(\d{2})', s)
                if match:
                    year = '20' + match.group(1)
                    return f'{year}-{month_num}-01'
        return None

    print("Pulizia e standardizzazione delle date...")
    df_merged['Mese_dt'] = pd.to_datetime(df_merged['Mese'].apply(clean_and_parse_date), errors='coerce')
    df_merged.dropna(subset=['Mese_dt'], inplace=True)

    if df_merged.empty:
        print("ATTENZIONE: Nessuna data valida trovata dopo la pulizia. Il report sarà vuoto.")
        # df_final = pd.DataFrame(columns=['Mese']) # This line was causing issues, removed
        # Create an empty workbook and save it if no data
        wb = Workbook()
        ws = wb.active
        ws.title = "Report Macchine"
        wb.save(file_output)
    else:
        df_merged['AnnoMese'] = df_merged['Mese_dt'].dt.strftime('%Y-%m')
        df_monthly = df_merged.groupby(['AnnoMese', 'Macchina']).agg({
            'Quantita Prodotta': 'sum',
            'Consumo Energia': 'sum'
        }).reset_index()
        df_monthly.rename(columns={'AnnoMese': 'Mese'}, inplace=True)

        # --- ANALISI ANOMALIE ---
        print("Eseguo l'analisi delle anomalie sull'efficienza...")
        df_anomalies = df_monthly.copy()
        df_anomalies['Consumo_per_Pezzo'] = df_anomalies['Consumo Energia'].divide(df_anomalies['Quantita Prodotta']).fillna(0)
        df_anomalies.replace([float('inf'), float('-inf')], 0, inplace=True)
        efficiency_series = df_anomalies[df_anomalies['Consumo_per_Pezzo'] > 0]['Consumo_per_Pezzo']
        if not efficiency_series.empty:
            mean_efficiency = efficiency_series.mean()
            std_efficiency = efficiency_series.std()
            anomaly_threshold_high = mean_efficiency + (2 * std_efficiency)
            anomaly_threshold_low = max(0, mean_efficiency - (2 * std_efficiency))
            print(f"Soglia anomalia alta (consumo/pezzo): > {anomaly_threshold_high:.2f}")
            anomalous_rows = df_anomalies[
                (df_anomalies['Consumo_per_Pezzo'] > anomaly_threshold_high) |
                ((df_anomalies['Consumo_per_Pezzo'] < anomaly_threshold_low) & (df_anomalies['Consumo_per_Pezzo'] > 0))
            ].copy()
            if not anomalous_rows.empty:
                def get_anomaly_reason(row):
                    if row['Consumo_per_Pezzo'] > anomaly_threshold_high:
                        return f"Consumo/Pezzo ALTO ({row['Consumo_per_Pezzo']:.2f} vs media {mean_efficiency:.2f})"
                    elif row['Consumo_per_Pezzo'] < anomaly_threshold_low:
                        return f"Consumo/Pezzo BASSO ({row['Consumo_per_Pezzo']:.2f} vs media {mean_efficiency:.2f})"
                    return "N/A"
                anomalous_rows.loc[:, 'Motivo_Anomalia'] = anomalous_rows.apply(get_anomaly_reason, axis=1)
                print(f"Trovate {len(anomalous_rows)} anomalie. Salvo in report_anomalie.xlsx")
                anomalous_rows.to_excel('report_anomalie.xlsx', index=False)
            else:
                print("Nessuna anomalia di efficienza trovata.")
        else:
            print("Non è stato possibile calcolare l'efficienza (nessuna produzione > 0). Salto l'analisi anomalie.")

        # --- PREPARAZIONE PER IL FORMATO "SPLITTATO" SU RIGHE ---
        print("Preparo il report per il formato 'splittato' su righe...")
        machines_per_row = 3 # Numero di macchine per riga
        
        all_machines = df_monthly['Macchina'].unique().tolist()
        all_machines.sort() # Ordino per consistenza
        
        # Initialize openpyxl workbook and worksheet
        # from openpyxl import Workbook # Already imported at top
        # from openpyxl.utils import get_column_letter # Already imported at top
        # from openpyxl.styles import Alignment # Already imported at top
        # from openpyxl.worksheet.properties import PageSetup, PrintPageSetup # Already imported at top

        wb = Workbook()
        ws = wb.active
        ws.title = "Report Macchine"

        # Set Page Layout (A4 Vertical)
        ws.page_setup.orientation = 'portrait' # Use string value
        ws.page_setup.paperSize = 9 # A4 paper size code
        print("Impostato layout di stampa: A4 Verticale.")

        current_row = 1 # Start writing from row 1 in Excel

        # Iterate through machines in chunks
        for i, group_start_idx in enumerate(range(0, len(all_machines), machines_per_row)):
            current_machines_chunk = all_machines[group_start_idx:group_start_idx + machines_per_row]
            
            # Filter df_monthly for the current chunk of machines
            df_chunk_monthly = df_monthly[df_monthly['Macchina'].isin(current_machines_chunk)]
            
            # Pivot this chunk's data
            df_chunk_wide = df_chunk_monthly.pivot(index='Mese', columns='Macchina', values=['Quantita Prodotta', 'Consumo Energia'])
            df_chunk_wide = df_chunk_wide.swaplevel(0, 1, axis=1)
            df_chunk_wide.sort_index(axis=1, level=[0, 1], inplace=True)
            
            # --- Write Headers for the current chunk ---
            # First header row (Machine names)
            ws.cell(row=current_row, column=1, value="Mese") # 'Mese' column header
            current_col = 2 # Start from column B
            for machine_name in current_machines_chunk:
                ws.cell(row=current_row, column=current_col, value=machine_name)
                # Merge cells for machine name (e.g., F01 spans 2 columns: Quantita, Consumo)
                ws.merge_cells(start_row=current_row, start_column=current_col, end_row=current_row, end_column=current_col + 1)
                current_col += 2 # Move to the next machine's start column

            current_row += 1 # Move to the second header row

            # Second header row (Metrics)
            ws.cell(row=current_row, column=1, value="") # Empty cell for Mese column
            current_col = 2 # Start from column B for metrics
            for machine_name in current_machines_chunk:
                ws.cell(row=current_row, column=current_col, value="kWh")
                ws.cell(row=current_row, column=current_col + 1, value="Pezzi")
                current_col += 2

            current_row += 1 # Move to data rows

            # --- Write Data Rows for the current chunk ---
            # Convert DataFrame to list of lists for writing
            data_to_write = df_chunk_wide.reset_index().values.tolist()
            for row_data in data_to_write:
                ws.append(row_data)
                current_row += 1
            
            # Add a separator (empty row) after each chunk, unless it's the last one
            if group_start_idx + machines_per_row < len(all_machines):
                ws.append([]) # Add an empty row
                current_row += 1 # Increment row counter for the empty row

        # --- Applica bordi a tutte le celle usate ---
        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
        print("Bordi applicati a tutte le celle.")

        # --- Regola la larghezza delle colonne (semplificato) ---
        default_width = 15
        for col_idx in range(1, ws.max_column + 1):
            column_letter = get_column_letter(col_idx)
            ws.column_dimensions[column_letter].width = default_width
        print(f"Impostata larghezza colonne predefinita: {default_width}.")

        # --- SALVATAGGIO DEL WORKBOOK ---
        print(f"Salvo il report finale formattato in: {file_output}")
        wb.save(file_output)
        print(f"\nOperazione completata! Controlla il file '{file_output}'.")

    # --- REPORT GLOBALE MENSILE ---
    print("\nGenerazione del report globale mensile...")

    # 1. Creazione del riassunto globale da df_monthly
    # df_monthly è già aggregato per Mese e Macchina. Ora aggrego per solo Mese.
    df_global_summary = df_monthly.groupby('Mese').agg({
        'Quantita Prodotta': 'sum',
        'Consumo Energia': 'sum'
    }).reset_index()

    df_global_summary.rename(columns={
        'Consumo Energia': 'Consumo Globale kWh',
        'Quantita Prodotta': 'Pezzi Prodotti Totali'
    }, inplace=True)

    # 2. Lettura e preparazione del file bollette.xlsx
    file_bollette = 'bollette.xlsx'
    if os.path.exists(file_bollette):
        print(f"Leggo il file: {file_bollette}")
        # Leggo il file senza intestazioni, poi seleziono le colonne per indice
        df_bollette = pd.read_excel(file_bollette, header=None) 
        
        # Seleziono la prima colonna (indice 0) come 'Mese' e l'ultima colonna (indice -1) come 'Valore Bolletta'
        df_bollette = df_bollette[[0, df_bollette.shape[1] - 1]] # Select first and last column
        df_bollette.columns = ['Mese', 'Valore Bolletta'] # Assign new column names

        # Assicuro che la colonna 'Mese' sia nello stesso formato 'YYYY-MM'
        df_bollette['Mese'] = pd.to_datetime(df_bollette['Mese'], errors='coerce').dt.strftime('%Y-%m')
        df_bollette.dropna(subset=['Mese'], inplace=True) # Rimuovo righe con mesi non validi

        # 3. Unione con il riassunto globale
        df_global_summary = pd.merge(
            df_global_summary,
            df_bollette,
            on='Mese',
            how='left' # Uso left merge per mantenere tutti i mesi del report globale
        )
        print("Valori bolletta integrati.")
    else:
        print(f"ATTENZIONE: Il file '{file_bollette}' non trovato. Il report globale non includerà il valore della bolletta.")
        df_global_summary['Valore Bolletta'] = None # Aggiungo la colonna ma la lascio vuota

    # 4. Salvataggio del report globale
    output_global_file = 'report_globale_mensile.xlsx'
    df_global_summary.to_excel(output_global_file, index=False)
    print(f"Report globale mensile salvato in: {output_global_file}")

except Exception as e:
    print(f"Si è verificato un errore: {e}")
    print("Assicurati di avere le librerie 'pandas' e 'openpyxl' installate.")
    print("Puoi installarle con il comando: pip install pandas openpyxl")

