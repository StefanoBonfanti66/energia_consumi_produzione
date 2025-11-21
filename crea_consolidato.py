
import pandas as pd
import openpyxl
import os
import yaml
import logging
from typing import Optional
from pydantic import BaseModel, ValidationError, field_validator

# --- Configurazione del Logging ---
logging.basicConfig(
    filename='errori_validazione.log',
    level=logging.ERROR,
    format='%(asctime)s - FOGLIO: %(sheet_name)s - RIGA: %(row_num)s - ERRORE: %(message)s',
    filemode='w'
)
logger = logging.getLogger()

# --- Modello di Dati con Pydantic ---
class DatiMacchinaRow(BaseModel):
    """Definisce la struttura e i tipi di una riga di dati valida."""
    macchina_o_impianto: str
    anno: int
    mese: int
    ore_produzione_macchina: Optional[float] = None
    pezzi_prodotti: Optional[float] = None
    consumo: Optional[float] = None
    lettura: Optional[float] = None
    costo_energia: Optional[float] = None
    costo_macchina: Optional[float] = None
    consumo_da_bolletta: Optional[float] = None
    totale_bolletta: Optional[float] = None

    @field_validator('*', mode='before')
    def empty_str_to_none(cls, v):
        if isinstance(v, str) and v.strip() in ('', '-'):
            return None
        return v

def load_config(config_path='config.yaml'):
    """Carica la configurazione dal file YAML."""
    if not os.path.exists(config_path):
        raise FileNotFoundError(f"File di configurazione non trovato: {config_path}")
    with open(config_path, 'r') as f:
        try:
            return yaml.safe_load(f)
        except yaml.YAMLError as e:
            raise ValueError(f"Errore nel parsing del file di configurazione YAML: {e}")

def crea_foglio_consolidato():
    """
    Legge, valida e consolida i dati, salvando gli errori in un file di log.
    """
    try:
        config = load_config()
        file_path = config['file_excel']
        sheets_to_exclude = set(config['fogli_da_escludere'])
        column_mapping = config['mappatura_colonne']
        final_columns = config['colonne_finali']
    except (FileNotFoundError, ValueError, KeyError) as e:
        print(f"Errore di configurazione: {e}")
        return

    if not os.path.exists(file_path):
        print(f"Errore: Il file '{file_path}' non è stato trovato.")
        return

    try:
        xls = pd.ExcelFile(file_path, engine='openpyxl')
        sheet_names = [s for s in xls.sheet_names if s not in sheets_to_exclude]
        print(f"Fogli trovati da elaborare: {sheet_names}")

        if not sheet_names:
            print("Nessun foglio di macchina trovato da elaborare.")
            return

        valid_rows = []
        error_count = 0

        for sheet_name in sheet_names:
            print(f"Leggo e valido il foglio: {sheet_name}...")
            df = pd.read_excel(xls, sheet_name=sheet_name, header=0)
            
            # Standardizza i nomi delle colonne per la validazione
            df.columns = df.columns.str.strip().str.replace(' ', '_')

            for index, row in df.iterrows():
                if pd.isna(row.get('macchina_o_impianto')) or not str(row.get('macchina_o_impianto')).strip():
                    continue
                
                try:
                    validated_row = DatiMacchinaRow(**row.to_dict())
                    valid_rows.append(validated_row.model_dump())
                except ValidationError as e:
                    error_details = e.errors()[0]
                    err_msg = f"{error_details['msg']} (colonna: {error_details['loc'][0]})"
                    logger.error(err_msg, extra={'sheet_name': sheet_name, 'row_num': index + 2})
                    error_count += 1
        
        if error_count > 0:
            print(f"ATTENZIONE: Sono stati trovati {error_count} errori di validazione. Controlla il file 'errori_validazione.log'.")

        if not valid_rows:
            print("Nessun dato valido trovato nei fogli delle macchine.")
            return

        consolidated_df = pd.DataFrame(valid_rows)
        print("Dati validi uniti con successo.")

        # Rinomina le colonne per il formato finale
        # Le colonne nel df sono es: 'macchina_o_impianto', nella config la key è 'macchina o impianto'
        inverted_mapping = {k.replace(' ', '_'): v for k, v in column_mapping.items()}
        consolidated_df.rename(columns=inverted_mapping, inplace=True)
        
        for col in final_columns:
            if col not in consolidated_df.columns:
                consolidated_df[col] = None
        
        consolidated_df = consolidated_df[final_columns]

        print(f"Salvataggio del foglio 'Consolidato' nel file '{file_path}'...")
        try:
            from openpyxl.utils.dataframe import dataframe_to_rows
            book = openpyxl.load_workbook(file_path)
            if 'Consolidato' in book.sheetnames:
                book.remove(book['Consolidato'])
            new_sheet = book.create_sheet('Consolidato')
            for r in dataframe_to_rows(consolidated_df, index=False, header=True):
                new_sheet.append(r)
            book.save(file_path)
            print("Operazione completata con successo!")

        except Exception as e:
            print(f"Errore durante il salvataggio con openpyxl: {e}")
            print("Tentativo di salvataggio alternativo con Pandas...")
            try:
                with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    consolidated_df.to_excel(writer, sheet_name='Consolidato', index=False)
                print("Salvataggio alternativo con Pandas riuscito!")
            except Exception as e2:
                print(f"Anche il salvataggio alternativo è fallito: {e2}")

    except Exception as e:
        print(f"Si è verificato un errore imprevisto: {e}")

if __name__ == "__main__":
    crea_foglio_consolidato()
